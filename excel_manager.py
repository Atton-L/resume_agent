"""
Excel表格管理模块
负责应聘者数据的Excel文件操作
"""
import os
from datetime import datetime
from typing import List, Optional, Dict, Any
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Excel文件路径
EXCEL_FILE = os.path.join(os.path.dirname(__file__), "data", "candidates.xlsx")

# 表格列定义
COLUMNS = [
    "序号",
    "姓名",
    "简历附件",
    "方向",
    "简历上传日期",
    "上传人",
    "工作base",
    "是否可以约面",
    "约面负责人",
    "初面时间",
    "面试官",
    "初面评价",
    "初面结论",
    "招聘状态"
]


def init_excel():
    """初始化Excel文件，创建表头；如果已存在则迁移列结构"""
    if not os.path.exists(EXCEL_FILE):
        # 确保目录存在
        os.makedirs(os.path.dirname(EXCEL_FILE), exist_ok=True)

        # 创建空DataFrame并保存
        df = pd.DataFrame(columns=COLUMNS)
        df.to_excel(EXCEL_FILE, index=False, engine='openpyxl')

        # 设置样式
        _format_excel()
        print(f"Excel文件已创建: {EXCEL_FILE}")
    else:
        # 检查已有文件的列是否需要迁移
        df = pd.read_excel(EXCEL_FILE, engine='openpyxl')
        existing_cols = list(df.columns)
        need_save = False

        # 添加缺失的新列
        for col in COLUMNS:
            if col not in existing_cols:
                df[col] = ''
                need_save = True

        # 删除不再需要的旧列
        old_cols = [c for c in existing_cols if c not in COLUMNS]
        if old_cols:
            df = df.drop(columns=old_cols)
            need_save = True

        # 按照 COLUMNS 顺序重排列
        if need_save:
            df = df[COLUMNS]
            df.to_excel(EXCEL_FILE, index=False, engine='openpyxl')
            _format_excel()
            print(f"Excel文件已迁移列结构: {EXCEL_FILE}")


def _format_excel():
    """格式化Excel表格样式"""
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    # 设置表头样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for col_num, column_title in enumerate(ws[1], 1):
        column_title.fill = header_fill
        column_title.font = header_font
        column_title.alignment = Alignment(horizontal='center', vertical='center')

    # 设置列宽
    column_widths = {
        'A': 8,   # 序号
        'B': 12,  # 姓名
        'C': 30,  # 简历附件
        'D': 10,  # 方向
        'E': 12,  # 简历上传日期
        'F': 10,  # 上传人
        'G': 10,  # 工作base
        'H': 12,  # 是否可以约面
        'I': 12,  # 约面负责人
        'J': 16,  # 初面时间
        'K': 10,  # 面试官
        'L': 16,  # 初面评价
        'M': 12,  # 初面结论
        'N': 12   # 招聘状态
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    wb.save(EXCEL_FILE)


def get_all_candidates() -> List[Dict[str, Any]]:
    """获取所有应聘者数据"""
    if not os.path.exists(EXCEL_FILE):
        init_excel()

    df = pd.read_excel(EXCEL_FILE, engine='openpyxl')
    df = df.fillna('')  # 将NaN转换为空字符串
    return df.to_dict('records')


def add_candidate(candidate_data: Dict[str, Any]) -> Dict[str, Any]:
    """添加新应聘者"""
    if not os.path.exists(EXCEL_FILE):
        init_excel()

    # 读取现有数据
    df = pd.read_excel(EXCEL_FILE, engine='openpyxl')

    # 生成新序号
    if len(df) > 0:
        max_id = df['序号'].max()
        new_id = max_id + 1 if pd.notna(max_id) else 1
    else:
        new_id = 1

    # 构建新记录
    new_record = {
        '序号': new_id,
        '姓名': candidate_data.get('name', ''),
        '简历附件': candidate_data.get('resume_file', ''),
        '方向': candidate_data.get('direction', ''),
        '简历上传日期': candidate_data.get('upload_date', datetime.now().strftime('%Y-%m-%d')),
        '上传人': candidate_data.get('uploader', '系统'),
        '工作base': candidate_data.get('work_base', ''),
        '是否可以约面': candidate_data.get('can_interview', ''),
        '约面负责人': candidate_data.get('interview_owner', ''),
        '初面时间': candidate_data.get('interview_date', ''),
        '面试官': candidate_data.get('interviewer', ''),
        '初面评价': candidate_data.get('first_interview_review', ''),
        '初面结论': candidate_data.get('first_interview_conclusion', ''),
        '招聘状态': candidate_data.get('recruitment_status', '')
    }

    # 添加到DataFrame
    df = pd.concat([df, pd.DataFrame([new_record])], ignore_index=True)

    # 保存
    df.to_excel(EXCEL_FILE, index=False, engine='openpyxl')
    _format_excel()

    return new_record


def update_candidate(candidate_id: int, update_data: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    """更新应聘者信息"""
    if not os.path.exists(EXCEL_FILE):
        return None

    df = pd.read_excel(EXCEL_FILE, engine='openpyxl')

    # 查找记录
    idx = df[df['序号'] == candidate_id].index
    if len(idx) == 0:
        return None

    # 更新字段
    for key, value in update_data.items():
        if key in COLUMNS:
            df.loc[idx[0], key] = value

    # 保存
    df.to_excel(EXCEL_FILE, index=False, engine='openpyxl')
    _format_excel()

    # 返回更新后的记录
    row = df.loc[idx[0]].fillna('')
    return row.to_dict()


def delete_candidate(candidate_id: int) -> bool:
    """删除应聘者"""
    if not os.path.exists(EXCEL_FILE):
        return False

    df = pd.read_excel(EXCEL_FILE, engine='openpyxl')

    # 查找并删除记录
    idx = df[df['序号'] == candidate_id].index
    if len(idx) == 0:
        return False

    df = df.drop(idx[0])

    # 保存
    df.to_excel(EXCEL_FILE, index=False, engine='openpyxl')
    _format_excel()

    return True


def get_candidate(candidate_id: int) -> Optional[Dict[str, Any]]:
    """获取单个应聘者信息"""
    if not os.path.exists(EXCEL_FILE):
        return None

    df = pd.read_excel(EXCEL_FILE, engine='openpyxl')
    df = df.fillna('')

    result = df[df['序号'] == candidate_id]
    if len(result) == 0:
        return None

    return result.iloc[0].to_dict()


# 初始化模块
init_excel()
